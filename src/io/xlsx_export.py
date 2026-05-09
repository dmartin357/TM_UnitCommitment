"""
Shared Excel export for UC solution comparison (benchmark vs heuristic).

Accepts pre-processed dispatch arrays indexed by display period (0-based,
length = n_periods) for both thermal and renewable generators, plus optional
summary statistics.  Callers are responsible for computing these arrays from
their native data format (CBC JSON or heuristic Stage 2 result).

Output workbook
---------------
  "Dispatch"      : full generator × period matrix + per-period summary rows
  "Chart Data"    : aggregated generation by broad fuel type (for Generation chart)
  "Generation"    : stacked area chart — broad fuel (Nuclear / Coal / NG / Oil / …)
  "Cat Chart Data": aggregated generation by granular category (for Category Mix chart)
  "Category Mix"  : stacked area chart — granular category (Gas CC / Gas CT / Oil CT / …)
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.chart import AreaChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Fuel classification ───────────────────────────────────────────────────────

# Broad fuel types — Gas CC and Gas CT are both aggregated as "NG"
FUEL_ORDER = ["Nuclear", "Coal", "NG", "Oil", "Hydro", "Solar", "Wind", "Other"]

FUEL_COLORS = {
    "Nuclear": "8B1A1A",   # dark maroon
    "Coal":    "4E3524",   # dark brown
    "NG":      "C8A882",   # tan (natural gas, any type)
    "Oil":     "A0522D",   # sienna
    "Hydro":   "6BAED6",   # light blue
    "Solar":   "FDD835",   # yellow
    "Wind":    "2CA02C",   # green
    "Other":   "AAAAAA",   # gray
}

# Lighter fill versions for broad-fuel rows (renewables in dispatch sheet)
FUEL_ROW_FILL = {
    "Nuclear": "F5DCDC",
    "Coal":    "E8D5C4",
    "NG":      "F5EEE0",
    "Oil":     "EDD5C0",
    "Hydro":   "D9EEF8",
    "Solar":   "FFFCE0",
    "Wind":    "D8F0D8",
    "Other":   "EEEEEE",
}

# Maps the fuel_category field (gen.csv Column F) → broad fuel label for Generation chart
_CATEGORY_TO_FUEL: dict[str, str] = {
    "Nuclear": "Nuclear",
    "Coal":    "Coal",
    "Gas CC":  "NG",
    "Gas CT":  "NG",
    "Oil CT":  "Oil",
    "Oil ST":  "Oil",
}

# Granular category order/colors for the Category Mix chart
CATEGORY_ORDER = [
    "Nuclear", "Coal", "Gas CC", "Gas CT", "Oil CT", "Oil ST",
    "Hydro", "Solar", "Wind", "Other",
]

CATEGORY_COLORS = {
    "Nuclear": "8B1A1A",
    "Coal":    "4E3524",
    "Gas CC":  "C8A882",   # tan
    "Gas CT":  "E8C99A",   # light tan
    "Oil CT":  "A0522D",   # sienna
    "Oil ST":  "C4763A",   # lighter sienna
    "Hydro":   "6BAED6",
    "Solar":   "FDD835",
    "Wind":    "2CA02C",
    "Other":   "AAAAAA",
}

# Lighter fill versions for granular-category rows (thermals in dispatch sheet)
CATEGORY_ROW_FILL = {
    "Nuclear": "F5DCDC",
    "Coal":    "E8D5C4",
    "Gas CC":  "F5EEE0",
    "Gas CT":  "FAF4E8",
    "Oil CT":  "EDD5C0",
    "Oil ST":  "F5E5D0",
    "Hydro":   "D9EEF8",
    "Solar":   "FFFCE0",
    "Wind":    "D8F0D8",
    "Other":   "EEEEEE",
}


def classify_fuel(name: str, gen_data: dict | None = None) -> str:
    """Return the broad fuel label (Nuclear / Coal / NG / Oil / Hydro / Solar / Wind)."""
    if gen_data is not None:
        cat = gen_data.get("fuel_category")
        if cat and cat in _CATEGORY_TO_FUEL:
            return _CATEGORY_TO_FUEL[cat]
    n = name.upper()
    if "NUCLEAR" in n:
        return "Nuclear"
    if "HYDRO" in n:
        return "Hydro"
    if "WIND" in n:
        return "Wind"
    if re.search(r"PV|CSP|SOLAR|RTPV", n):
        return "Solar"
    if re.search(r"STEAM|COAL", n):
        return "Coal"
    if "CC" in n or re.search(r"\bCT\b|\bGT\b|GAS", n):
        return "NG"
    return "Other"


# Maps fuel_category (gen.csv Column F) → granular category label
_CATEGORY_TO_CATEGORY: dict[str, str] = {
    "Nuclear": "Nuclear",
    "Coal":    "Coal",
    "Gas CC":  "Gas CC",
    "Gas CT":  "Gas CT",
    "Oil CT":  "Oil CT",
    "Oil ST":  "Oil ST",
}


def classify_category(name: str, gen_data: dict | None = None) -> str:
    """Return the granular category label (Gas CC / Gas CT / Oil CT / Oil ST / …)."""
    if gen_data is not None:
        cat = gen_data.get("fuel_category")
        if cat and cat in _CATEGORY_TO_CATEGORY:
            return _CATEGORY_TO_CATEGORY[cat]
    n = name.upper()
    if "NUCLEAR" in n:
        return "Nuclear"
    if "HYDRO" in n:
        return "Hydro"
    if "WIND" in n:
        return "Wind"
    if re.search(r"PV|CSP|SOLAR|RTPV", n):
        return "Solar"
    if re.search(r"STEAM|COAL", n):
        return "Coal"
    if "CC" in n:
        return "Gas CC"
    if re.search(r"\bCT\b|\bGT\b|GAS", n):
        return "Gas CT"
    return "Other"


# ── Style constants ───────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill(fill_type="solid", fgColor="2E4057")   # dark blue
_HDR_FONT  = Font(bold=True, color="FFFFFF")
_SECT_FILL = PatternFill(fill_type="solid", fgColor="C5D1E0")   # section separator
_SUMM_FILL = PatternFill(fill_type="solid", fgColor="EEF2F7")   # summary rows
_BOLD      = Font(bold=True)
_CTR       = Alignment(horizontal="center")
_RIGHT     = Alignment(horizontal="right")


def _auto_size_columns(ws, padding: int = 2, min_width: int = 5) -> None:
    """Set each column's width to fit its widest cell content.

    Currency-formatted cells are rendered as '$1,234,567.89' for measurement
    since openpyxl stores the raw float but Excel displays the formatted string.
    Merged cells (value=None on subordinate cells) are skipped.
    """
    for col_cells in ws.columns:
        max_width = min_width
        for cell in col_cells:
            if cell.value is None:
                continue
            fmt = cell.number_format or ""
            if "$" in fmt:
                try:
                    rendered = f"${abs(float(cell.value)):,.2f}"
                except (TypeError, ValueError):
                    rendered = str(cell.value)
            else:
                rendered = str(cell.value)
            max_width = max(max_width, len(rendered))
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max_width + padding


# ── Main export function ──────────────────────────────────────────────────────

def export_solution_xlsx(
    path: Path,
    generators_thermal: dict,
    generators_renewable: dict,
    n_periods: int,
    thermal_dispatch: dict[str, list[float | None]],
    renewable_dispatch: dict[str, list[float | None]],
    thermal_demand: list[float] | None = None,
    renewable_expected: list[float] | None = None,
    renewable_min_vals: list[float] | None = None,
    renewable_max_vals: list[float] | None = None,
    committed_thermal: list[int | None] | None = None,
    reg_up_total: list[float | None] | None = None,
    reg_down_total: list[float | None] | None = None,
    spin_up_total: list[float | None] | None = None,
    spin_down_total: list[float | None] | None = None,
    flex_up_total: list[float | None] | None = None,
    flex_down_total: list[float | None] | None = None,
    ramp_up_total: list[float | None] | None = None,
    ramp_down_total: list[float | None] | None = None,
    ed_cost_per_period: list[float | None] | None = None,
    startup_cost_per_period: list[float | None] | None = None,
    period_labels: list[str] | None = None,
    sheet_title: str = "Dispatch",
) -> Path:
    """
    Write a solution comparison xlsx.

    Parameters
    ----------
    path                 : Output file path.
    generators_thermal   : {name: gen_data} thermal generators from instance JSON.
    generators_renewable : {name: gen_data} renewable generators from instance JSON.
    n_periods            : Number of time periods.
    thermal_dispatch     : {gen_name: [mw_t0, ..., mw_tN-1]} — total MW per period.
                           Use None entries for periods where the unit is off or unknown.
    renewable_dispatch   : Same structure for renewable (including hydro) generators.
    thermal_demand       : Thermal demand (MW) per period.
    renewable_expected   : Expected renewable output (MW) per period.
    renewable_min_vals   : Min renewable output (MW) per period.
    renewable_max_vals   : Max renewable output (MW) per period.
    committed_thermal    : Number of committed thermal generators per period.
    reg_up_total         : Total regulation-up available (MW) per period.
    reg_down_total       : Total regulation-down available (MW) per period.
    spin_up_total        : Total spinning reserve up (MW) per period.
    spin_down_total      : Total spinning reserve down (MW) per period.
    flex_up_total        : Total 20-min flex up (MW) per period.
    flex_down_total      : Total 20-min flex down (MW) per period.
    ramp_up_total        : Total hourly ramp up (MW) per period.
    ramp_down_total      : Total hourly ramp down (MW) per period.
    ed_cost_per_period   : Economic dispatch (thermal) cost per period ($).
    startup_cost_per_period : Startup cost per period ($).
    period_labels        : Column header labels ["t=1", ..., "t=48"].
                           Defaults to ["t=1", ..., "t=N"].
    sheet_title          : Name of the main dispatch sheet.
    """
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if period_labels is None:
        period_labels = [f"t={t+1}" for t in range(n_periods)]

    # ── Sort generators ───────────────────────────────────────────────────────

    thermal_names = sorted(
        generators_thermal.keys(),
        key=lambda n: generators_thermal[n].get("power_output_maximum", 0.0),
        reverse=True,
    )

    # Renewables: hydro first, then solar, then wind, then other
    def _ren_sort_key(n):
        fuel = classify_fuel(n, generators_renewable.get(n))
        return (FUEL_ORDER.index(fuel) if fuel in FUEL_ORDER else 99,
                -generators_renewable[n].get("power_output_maximum", [0.0])[0]
                if isinstance(generators_renewable[n].get("power_output_maximum"), list)
                else -generators_renewable[n].get("power_output_maximum", 0.0))

    renewable_names = sorted(generators_renewable.keys(), key=_ren_sort_key)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Dispatch ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = sheet_title

    DATA_COL_START = 5   # columns A=name B=fuel C=pmax D=pmin E+=periods

    # Header row
    for col, val in enumerate(["Unit", "Fuel", "Pmax (MW)", "Pmin (MW)"], start=1):
        c = ws.cell(row=1, column=col, value=val)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = _CTR

    for t_idx, label in enumerate(period_labels):
        col = DATA_COL_START + t_idx
        c = ws.cell(row=1, column=col, value=label)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = _CTR

    # ── Thermal rows ──────────────────────────────────────────────────────────
    thermal_start_row = 2
    for row_offset, name in enumerate(thermal_names):
        row = thermal_start_row + row_offset
        gen      = generators_thermal[name]
        category = classify_category(name, gen)   # granular: Gas CC, Gas CT, Oil CT …
        pmax = gen.get("power_output_maximum", 0.0)
        pmin = gen.get("power_output_minimum", 0.0)

        fill = PatternFill(fill_type="solid",
                           fgColor=CATEGORY_ROW_FILL.get(category, "EEEEEE"))
        name_cell = ws.cell(row=row, column=1, value=name)
        name_cell.font = _BOLD
        name_cell.fill = fill
        ws.cell(row=row, column=2, value=category).fill = fill
        ws.cell(row=row, column=3, value=round(pmax, 1)).alignment = _RIGHT
        ws.cell(row=row, column=4, value=round(pmin, 1)).alignment = _RIGHT

        dispatch_row = thermal_dispatch.get(name, [None] * n_periods)
        for t_idx in range(n_periods):
            mw = dispatch_row[t_idx] if t_idx < len(dispatch_row) else None
            col = DATA_COL_START + t_idx
            c = ws.cell(row=row, column=col,
                        value=round(mw, 2) if mw is not None else 0.0)
            c.alignment = _RIGHT

    thermal_end_row = thermal_start_row + len(thermal_names) - 1

    # ── Renewable section header ──────────────────────────────────────────────
    sep_row = thermal_end_row + 1
    ws.row_dimensions[sep_row].height = 14
    total_cols = DATA_COL_START + n_periods - 1
    ws.merge_cells(start_row=sep_row, start_column=1,
                   end_row=sep_row, end_column=total_cols)
    sep_cell = ws.cell(row=sep_row, column=1, value="Renewables (Hydro / Solar / Wind)")
    sep_cell.font = Font(bold=True, color="2E4057")
    sep_cell.fill = _SECT_FILL
    sep_cell.alignment = _CTR

    # ── Renewable rows ────────────────────────────────────────────────────────
    ren_start_row = sep_row + 1
    for row_offset, name in enumerate(renewable_names):
        row = ren_start_row + row_offset
        gen  = generators_renewable[name]
        fuel = classify_fuel(name, gen)

        pmax_raw = gen.get("power_output_maximum", [0.0])
        pmin_raw = gen.get("power_output_minimum", [0.0])
        pmax_display = max(pmax_raw) if isinstance(pmax_raw, list) else float(pmax_raw)
        pmin_display = min(pmin_raw) if isinstance(pmin_raw, list) else float(pmin_raw)

        fill = PatternFill(fill_type="solid", fgColor=FUEL_ROW_FILL.get(fuel, "EEEEEE"))
        name_cell = ws.cell(row=row, column=1, value=name)
        name_cell.font = _BOLD
        name_cell.fill = fill
        for col in range(2, DATA_COL_START):
            ws.cell(row=row, column=col).fill = fill

        ws.cell(row=row, column=2, value=fuel).fill = fill
        ws.cell(row=row, column=3, value=round(pmax_display, 1)).alignment = _RIGHT
        ws.cell(row=row, column=4, value=round(pmin_display, 1)).alignment = _RIGHT

        dispatch_row = renewable_dispatch.get(name, [None] * n_periods)
        for t_idx in range(n_periods):
            mw = dispatch_row[t_idx] if t_idx < len(dispatch_row) else None
            col = DATA_COL_START + t_idx
            c = ws.cell(row=row, column=col,
                        value=round(mw, 2) if mw is not None else 0.0)
            c.alignment = _RIGHT
            c.fill = fill

    ren_end_row = ren_start_row + len(renewable_names) - 1

    # ── Summary section ───────────────────────────────────────────────────────
    sum_sep_row = ren_end_row + 1
    ws.merge_cells(start_row=sum_sep_row, start_column=1,
                   end_row=sum_sep_row, end_column=total_cols)
    sum_sep = ws.cell(row=sum_sep_row, column=1, value="Period Summary")
    sum_sep.font = Font(bold=True, color="2E4057")
    sum_sep.fill = _SECT_FILL
    sum_sep.alignment = _CTR

    total_cost_per_period: list[float | None] | None = None
    if ed_cost_per_period is not None or startup_cost_per_period is not None:
        ed  = ed_cost_per_period      or [0.0] * n_periods
        su  = startup_cost_per_period or [0.0] * n_periods
        total_cost_per_period = [
            (ed[t] or 0.0) + (su[t] or 0.0) for t in range(n_periods)
        ]

    _CCY = '$#,##0.00'   # currency format applied to cost rows

    # Each entry: (label, values, number_format)
    summary_rows = [
        ("Thermal Demand (MW)",         thermal_demand,         None),
        ("Renewable Expected (MW)",     renewable_expected,     None),
        ("Renewable Min (MW)",          renewable_min_vals,     None),
        ("Renewable Max (MW)",          renewable_max_vals,     None),
        ("Committed Thermal (#)",       committed_thermal,      None),
        ("ED Cost per Period ($)",      ed_cost_per_period,     _CCY),
        ("Startup Cost per Period ($)", startup_cost_per_period, _CCY),
        ("Total Cost per Period ($)",   total_cost_per_period,  _CCY),
        ("Reg Up Available (MW)",       reg_up_total,           None),
        ("Reg Down Available (MW)",     reg_down_total,         None),
        ("Spin Up Available (MW)",      spin_up_total,          None),
        ("Spin Down Available (MW)",    spin_down_total,        None),
        ("Flex Up Available (MW)",      flex_up_total,          None),
        ("Flex Down Available (MW)",    flex_down_total,        None),
    ]

    # Compute total dispatch per period
    total_dispatch = []
    for t_idx in range(n_periods):
        tot = 0.0
        for name in thermal_names:
            mw = thermal_dispatch.get(name, [None] * n_periods)
            v = mw[t_idx] if t_idx < len(mw) else None
            tot += (v or 0.0)
        for name in renewable_names:
            mw = renewable_dispatch.get(name, [None] * n_periods)
            v = mw[t_idx] if t_idx < len(mw) else None
            tot += (v or 0.0)
        total_dispatch.append(tot)
    summary_rows.append(("Total Dispatch (MW)", total_dispatch, None))

    for row_offset, (label, values, fmt) in enumerate(summary_rows):
        row = sum_sep_row + 1 + row_offset
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = _BOLD
        label_cell.fill = _SUMM_FILL
        for col in range(2, DATA_COL_START):
            ws.cell(row=row, column=col).fill = _SUMM_FILL

        for t_idx in range(n_periods):
            col = DATA_COL_START + t_idx
            c = ws.cell(row=row, column=col)
            c.fill = _SUMM_FILL
            if values is not None and t_idx < len(values):
                v = values[t_idx]
                if v is not None:
                    c.value = float(v) if fmt else round(float(v), 1)
                    if fmt:
                        c.number_format = fmt
            c.alignment = _RIGHT

    # ── Heat map on thermal dispatch cells ────────────────────────────────────
    first_data_col = get_column_letter(DATA_COL_START)
    last_data_col  = get_column_letter(DATA_COL_START + n_periods - 1)
    thermal_range  = f"{first_data_col}{thermal_start_row}:{last_data_col}{thermal_end_row}"
    ws.conditional_formatting.add(
        thermal_range,
        ColorScaleRule(
            start_type="num",      start_value=0,  start_color="FFFFFF",
            mid_type="percentile", mid_value=50,   mid_color="FFF176",
            end_type="max",                         end_color="1B5E20",
        ),
    )

    # ── Column widths & freeze ────────────────────────────────────────────────
    _auto_size_columns(ws)
    ws.freeze_panes = f"{get_column_letter(DATA_COL_START)}2"

    # ── Sheet 2: Chart Data ───────────────────────────────────────────────────
    ws_cd = wb.create_sheet("Chart Data")

    # Determine active fuel types (those with any non-zero dispatch)
    fuel_totals: dict[str, list[float]] = {f: [0.0] * n_periods for f in FUEL_ORDER}

    for name in thermal_names:
        fuel = classify_fuel(name, generators_thermal.get(name))
        row_data = thermal_dispatch.get(name, [None] * n_periods)
        for t_idx in range(n_periods):
            v = row_data[t_idx] if t_idx < len(row_data) else None
            fuel_totals[fuel][t_idx] += v or 0.0

    for name in renewable_names:
        fuel = classify_fuel(name, generators_renewable.get(name))
        row_data = renewable_dispatch.get(name, [None] * n_periods)
        for t_idx in range(n_periods):
            v = row_data[t_idx] if t_idx < len(row_data) else None
            fuel_totals[fuel][t_idx] += v or 0.0

    active_fuels = [f for f in FUEL_ORDER if any(v > 0.01 for v in fuel_totals[f])]

    # Write chart data header
    ws_cd.cell(row=1, column=1, value="Period").font = _BOLD
    for col_off, fuel in enumerate(active_fuels, start=2):
        ws_cd.cell(row=1, column=col_off, value=fuel).font = _BOLD

    for t_idx in range(n_periods):
        row = t_idx + 2
        ws_cd.cell(row=row, column=1, value=period_labels[t_idx])
        for col_off, fuel in enumerate(active_fuels, start=2):
            ws_cd.cell(row=row, column=col_off,
                       value=round(fuel_totals[fuel][t_idx], 2))

    # ── Sheet 3: Generation Chart ─────────────────────────────────────────────
    ws_ch = wb.create_sheet("Generation")

    chart = AreaChart()
    chart.grouping  = "stacked"
    chart.title     = f"Generation by Fuel Type — {sheet_title}"
    chart.y_axis.title = "Generation (MW)"
    chart.x_axis.title = "Period"
    chart.y_axis.numFmt = "#,##0"
    chart.width  = 30
    chart.height = 18

    # Add data series (all fuel columns at once)
    data_ref = Reference(
        ws_cd,
        min_col=2,
        max_col=1 + len(active_fuels),
        min_row=1,
        max_row=n_periods + 1,
    )
    chart.add_data(data_ref, titles_from_data=True)

    # Set x-axis categories
    cats = Reference(ws_cd, min_col=1, min_row=2, max_row=n_periods + 1)
    chart.set_categories(cats)

    # Apply fuel colors to series
    for i, fuel in enumerate(active_fuels):
        hex_color = FUEL_COLORS.get(fuel, "AAAAAA")
        s = chart.series[i]
        s.graphicalProperties.solidFill        = hex_color
        s.graphicalProperties.line.solidFill   = hex_color
        s.graphicalProperties.line.width       = 6350   # hairline (EMU)

    ws_ch.add_chart(chart, "A1")

    # ── Sheet 4: Category Chart Data ──────────────────────────────────────────
    ws_ccd = wb.create_sheet("Cat Chart Data")

    cat_totals: dict[str, list[float]] = {c: [0.0] * n_periods for c in CATEGORY_ORDER}

    for name in thermal_names:
        cat      = classify_category(name, generators_thermal.get(name))
        row_data = thermal_dispatch.get(name, [None] * n_periods)
        for t_idx in range(n_periods):
            v = row_data[t_idx] if t_idx < len(row_data) else None
            cat_totals[cat][t_idx] += v or 0.0

    for name in renewable_names:
        cat      = classify_category(name, generators_renewable.get(name))
        row_data = renewable_dispatch.get(name, [None] * n_periods)
        for t_idx in range(n_periods):
            v = row_data[t_idx] if t_idx < len(row_data) else None
            cat_totals[cat][t_idx] += v or 0.0

    active_cats = [c for c in CATEGORY_ORDER if any(v > 0.01 for v in cat_totals[c])]

    ws_ccd.cell(row=1, column=1, value="Period").font = _BOLD
    for col_off, cat in enumerate(active_cats, start=2):
        ws_ccd.cell(row=1, column=col_off, value=cat).font = _BOLD

    for t_idx in range(n_periods):
        row = t_idx + 2
        ws_ccd.cell(row=row, column=1, value=period_labels[t_idx])
        for col_off, cat in enumerate(active_cats, start=2):
            ws_ccd.cell(row=row, column=col_off,
                        value=round(cat_totals[cat][t_idx], 2))

    # ── Sheet 5: Category Mix Chart ───────────────────────────────────────────
    ws_cmix = wb.create_sheet("Category Mix")

    cat_chart = AreaChart()
    cat_chart.grouping  = "stacked"
    cat_chart.title     = f"Generation by Category — {sheet_title}"
    cat_chart.y_axis.title = "Generation (MW)"
    cat_chart.x_axis.title = "Period"
    cat_chart.y_axis.numFmt = "#,##0"
    cat_chart.width  = 30
    cat_chart.height = 18

    cat_data_ref = Reference(
        ws_ccd,
        min_col=2,
        max_col=1 + len(active_cats),
        min_row=1,
        max_row=n_periods + 1,
    )
    cat_chart.add_data(cat_data_ref, titles_from_data=True)

    cat_cats = Reference(ws_ccd, min_col=1, min_row=2, max_row=n_periods + 1)
    cat_chart.set_categories(cat_cats)

    for i, cat in enumerate(active_cats):
        hex_color = CATEGORY_COLORS.get(cat, "AAAAAA")
        s = cat_chart.series[i]
        s.graphicalProperties.solidFill      = hex_color
        s.graphicalProperties.line.solidFill = hex_color
        s.graphicalProperties.line.width     = 6350

    ws_cmix.add_chart(cat_chart, "A1")

    wb.save(path)
    return path


# ── Reserve helpers (usable by both CBC and heuristic callers) ────────────────

def compute_all_reserves_per_period(
    generators: dict,
    thermal_dispatch: dict[str, list[float | None]],
    committed_per_period: list[list[str] | None],
    n_periods: int,
) -> dict[str, list[float]]:
    """
    Compute all 8 reserve categories per period.

    For each committed unit the available reserve is:
      min(reserve_limit, physical_headroom)
    where headroom_up = pmax - dispatch  and  headroom_dn = dispatch - pmin.

    Reserve field names in the generator JSON:
      reg_up_limit / reg_down_limit   : regulation  (≈ flex / 4)
      spin_up_limit / spin_down_limit : spinning     (≈ flex / 2)
      flex_up_limit / flex_down_limit : 20-min flex  (original ramp_up/dn values)
      ramp_up_limit / ramp_down_limit : hourly ramp  (≈ flex × 3)

    Parameters
    ----------
    generators           : {name: gen_data} thermal generators.
    thermal_dispatch     : {name: [mw per period]} total MW.
    committed_per_period : List (len=n_periods) of committed generator name lists.
                           Use None for periods where committed set is unknown.

    Returns
    -------
    dict with keys:
      'reg_up', 'reg_down', 'spin_up', 'spin_down',
      'flex_up', 'flex_down', 'ramp_up', 'ramp_down'
    each a list of floats, length = n_periods.
    """
    keys = ("reg_up", "reg_down", "spin_up", "spin_down",
            "flex_up", "flex_down", "ramp_up", "ramp_down")
    result = {k: [0.0] * n_periods for k in keys}

    limit_fields = {
        "reg_up":   "reg_up_limit",
        "reg_down": "reg_down_limit",
        "spin_up":  "spin_up_limit",
        "spin_down":"spin_down_limit",
        "flex_up":  "flex_up_limit",
        "flex_down":"flex_down_limit",
        "ramp_up":  "ramp_up_limit",
        "ramp_down":"ramp_down_limit",
    }

    for t_idx in range(n_periods):
        committed = committed_per_period[t_idx] if committed_per_period else None
        if committed is None:
            continue
        for name in committed:
            gen = generators.get(name)
            if gen is None:
                continue
            dispatch_list = thermal_dispatch.get(name, [])
            mw   = dispatch_list[t_idx] if t_idx < len(dispatch_list) else 0.0
            pmax = float(gen.get("power_output_maximum", 0.0))
            pmin = float(gen.get("power_output_minimum", 0.0))
            hw_up = max(0.0, pmax - mw)
            hw_dn = max(0.0, mw - pmin)

            for key in ("reg_up", "spin_up", "flex_up", "ramp_up"):
                lim = float(gen.get(limit_fields[key], 0.0))
                result[key][t_idx] += min(lim, hw_up)
            for key in ("reg_down", "spin_down", "flex_down", "ramp_down"):
                lim = float(gen.get(limit_fields[key], 0.0))
                result[key][t_idx] += min(lim, hw_dn)

    return result


def _interpolate_pw_cost(pw: list[dict], mw: float) -> float:
    """Return the total generation cost ($) at dispatch level mw using piecewise_production."""
    if not pw or mw <= 0:
        return 0.0
    if mw <= pw[0]["mw"] + 1e-9:
        return float(pw[0]["cost"])
    for i in range(1, len(pw)):
        p0, c0 = float(pw[i - 1]["mw"]), float(pw[i - 1]["cost"])
        p1, c1 = float(pw[i]["mw"]),     float(pw[i]["cost"])
        if mw <= p1 + 1e-9:
            if p1 - p0 < 1e-9:
                return c1
            return c0 + (c1 - c0) * (mw - p0) / (p1 - p0)
    return float(pw[-1]["cost"])


def compute_dispatch_cost_per_period(
    generators: dict,
    thermal_dispatch: dict[str, list[float | None]],
    n_periods: int,
) -> list[float]:
    """
    Compute total thermal ED cost per period by interpolating piecewise_production.

    Parameters
    ----------
    generators       : {name: gen_data} thermal generators.
    thermal_dispatch : {name: [mw per period]} total MW dispatched.
    n_periods        : Number of periods.

    Returns
    -------
    List of floats (length = n_periods), each the summed generation cost
    across all committed units in that period (in the same $ units used by the
    piecewise_production cost breakpoints — typically $/half-hour for RTS-GMLC).
    """
    ed_cost = [0.0] * n_periods
    for name, gen in generators.items():
        pw = gen.get("piecewise_production", [])
        if not pw:
            continue
        dispatch_list = thermal_dispatch.get(name, [])
        for t_idx in range(n_periods):
            mw = dispatch_list[t_idx] if t_idx < len(dispatch_list) else None
            if mw and mw > 0:
                ed_cost[t_idx] += _interpolate_pw_cost(pw, mw)
    return ed_cost


def compute_reg_per_period(
    generators: dict,
    thermal_dispatch: dict[str, list[float | None]],
    committed_per_period: list[list[str] | None],
    n_periods: int,
) -> tuple[list[float], list[float]]:
    """
    Compute total regulation-up and regulation-down per period.

    Uses reg_up_limit / reg_down_limit fields (≈ flex / 4).
    For full reserve breakdown use compute_all_reserves_per_period().
    """
    reserves = compute_all_reserves_per_period(
        generators, thermal_dispatch, committed_per_period, n_periods
    )
    return reserves["reg_up"], reserves["reg_down"]
