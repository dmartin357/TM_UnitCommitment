"""
Smoke test for Stage 2 GA forward pass.

Loads saved Stage 1 results (JSON) and pre-solve targets, then runs the
Stage 2 sequential forward pass and prints the per-period decision table.

Requires a Stage 1 result file produced by smoke_test_stage1.py
(set STAGE1_RESULT_PATH below).

Usage (from repo root, with power-systems conda env active):
    python testing/smoke_test_stage2_ga.py
"""

import json
import logging
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

import numpy as np

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.io.stage1_io import load_stage1_result
from src.stage2_ga.config import Stage2Config
from src.stage2_ga.forward_pass import run_stage2_forward_pass

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(name)s  %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout,
    force=True,
)

# ── Paths ─────────────────────────────────────────────────────────────────────
PGLIB_UC_ROOT      = Path("C:/gitrepos/power-grid-lib/pglib-uc")
INSTANCE_PATH      = PGLIB_UC_ROOT / "rts_gmlc" / "2020-01-27.json"
CACHE_DIR          = Path(__file__).parent / "cache"
STAGE1_RESULT_PATH = CACHE_DIR / f"stage1_{INSTANCE_PATH.stem}.json"

# ── Config ────────────────────────────────────────────────────────────────────
STAGE2_CONFIG = Stage2Config(
    solver="auto",
    target_percentile=50.0,
    selection_mode="uniform",
    rng_seed=42,
)

# ── Instance loader (mirrors smoke_test_stage1.py) ────────────────────────────

class InstanceData:
    def __init__(self, thermal, total_demand, renewable_min, renewable_max):
        self.thermal = thermal
        self.total_demand = total_demand
        n = len(total_demand)
        renewable_max_eff = [min(renewable_max[t], total_demand[t]) for t in range(n)]
        self.renewable_expected = [
            (renewable_min[t] + renewable_max_eff[t]) / 2.0 for t in range(n)
        ]
        self.thermal_demand = [
            max(0.0, total_demand[t] - self.renewable_expected[t]) for t in range(n)
        ]
        self.renewable_reg_up   = [self.renewable_expected[t] - renewable_min[t] for t in range(n)]
        self.renewable_reg_down = [renewable_max_eff[t] - self.renewable_expected[t] for t in range(n)]


def load_instance(path: Path) -> InstanceData:
    with open(path) as f:
        data = json.load(f)
    thermal = data["thermal_generators"]
    demand_raw = data["demand"]
    if isinstance(demand_raw, list):
        total_demand = [float(d) for d in demand_raw]
    elif isinstance(demand_raw, dict):
        n = len(next(iter(demand_raw.values())))
        total_demand = [sum(bus[t] for bus in demand_raw.values()) for t in range(n)]
    else:
        raise ValueError(f"Unexpected demand format: {type(demand_raw)}")
    n_periods = len(total_demand)
    renewable_min = [0.0] * n_periods
    renewable_max = [0.0] * n_periods
    for name, gen_data in data.get("renewable_generators", {}).items():
        if re.search(r"HYDRO", name, re.IGNORECASE):
            continue
        for t in range(n_periods):
            pmin_s = gen_data.get("power_output_minimum", [])
            pmax_s = gen_data.get("power_output_maximum", [])
            if t < len(pmin_s):
                renewable_min[t] += float(pmin_s[t])
            if t < len(pmax_s):
                renewable_max[t] += float(pmax_s[t])
    return InstanceData(thermal, total_demand, renewable_min, renewable_max)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"Loading instance: {INSTANCE_PATH}")
    inst = load_instance(INSTANCE_PATH)
    n_periods = len(inst.total_demand)
    print(f"Instance: {len(inst.thermal)} thermal generators, {n_periods} periods\n",
          flush=True)

    # ── Load Stage 1 results ──────────────────────────────────────────────────
    print(f"Loading Stage 1 results: {STAGE1_RESULT_PATH}")
    stage1_result = load_stage1_result(STAGE1_RESULT_PATH)
    populations   = stage1_result.populations
    print(f"Stage 1: {len(populations)} populations loaded\n", flush=True)

    # ── Run Stage 2 forward pass ──────────────────────────────────────────────
    print("Running Stage 2 forward pass…\n", flush=True)
    rng = np.random.default_rng(STAGE2_CONFIG.rng_seed)

    result = run_stage2_forward_pass(
        populations=populations,
        generators=inst.thermal,
        thermal_demand_values=inst.thermal_demand,
        config=STAGE2_CONFIG,
        rng=rng,
    )

    result.print_summary()

    xlsx_path = Path(__file__).parent / "results" / f"stage2_dispatch_{INSTANCE_PATH.stem}.xlsx"
    export_dispatch_excel(result, inst.thermal, xlsx_path)


# ── Excel export ──────────────────────────────────────────────────────────────

def export_dispatch_excel(result, generators: dict, path: Path) -> None:
    """
    Write a dispatch matrix to Excel:
      rows    = generators (sorted by pmax descending)
      columns = time periods (t=0 … t=n_periods-1)
      cells   = power output MW (0.0 if unit uncommitted or period has no decision)

    Applies a white→yellow→green color-scale heat map across all dispatch cells.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    # Build period → {unit: MW} lookup from decisions
    dispatch_by_period: dict[int, dict[str, float]] = {}
    for d in result.decisions:
        dispatch_by_period[d.period] = d.dispatch

    # Determine full period range from the generator data
    # t=0 comes from power_output_t0; remaining periods from decisions
    sample_gen = next(iter(generators.values()))
    # n_periods: infer from max period in decisions + 1, or at least 1
    n_periods = max((d.period for d in result.decisions), default=0) + 1

    # Sort generators by pmax descending
    sorted_names = sorted(
        generators.keys(),
        key=lambda n: generators[n].get("power_output_maximum", 0.0),
        reverse=True,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dispatch"

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill(fill_type="solid", fgColor="2E4057")
    header_font = Font(bold=True, color="FFFFFF")

    ws.cell(row=1, column=1, value="Unit").font        = header_font
    ws.cell(row=1, column=1).fill                      = header_fill
    ws.cell(row=1, column=1).alignment                 = Alignment(horizontal="center")
    ws.cell(row=1, column=2, value="Pmax (MW)").font   = header_font
    ws.cell(row=1, column=2).fill                      = header_fill
    ws.cell(row=1, column=2).alignment                 = Alignment(horizontal="center")
    ws.cell(row=1, column=3, value="Pmin (MW)").font   = header_font
    ws.cell(row=1, column=3).fill                      = header_fill
    ws.cell(row=1, column=3).alignment                 = Alignment(horizontal="center")

    for t in range(n_periods):
        col = t + 4
        cell = ws.cell(row=1, column=col, value=f"t={t}")
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, name in enumerate(sorted_names, start=2):
        gen  = generators[name]
        pmax = gen.get("power_output_maximum", 0.0)
        pmin = gen.get("power_output_minimum", 0.0)

        name_cell = ws.cell(row=row_idx, column=1, value=name)
        name_cell.font = Font(bold=True)

        ws.cell(row=row_idx, column=2, value=round(pmax, 1))
        ws.cell(row=row_idx, column=3, value=round(pmin, 1))

        for t in range(n_periods):
            col = t + 4
            if t == 0:
                mw = float(gen.get("power_output_t0", 0.0))
            else:
                mw = dispatch_by_period.get(t, {}).get(name, 0.0)
            cell = ws.cell(row=row_idx, column=col, value=round(mw, 2))
            cell.alignment = Alignment(horizontal="right")

    # ── Color-scale heat map on dispatch cells ────────────────────────────────
    n_units      = len(sorted_names)
    first_data_col = get_column_letter(4)
    last_data_col  = get_column_letter(n_periods + 3)
    data_range     = f"{first_data_col}2:{last_data_col}{n_units + 1}"

    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type="num",   start_value=0,   start_color="FFFFFF",  # white  = 0 MW
            mid_type="percentile", mid_value=50,  mid_color="FFF176",   # yellow = median
            end_type="max",                        end_color="1B5E20",   # dark green = max
        ),
    )

    # ── Column widths & freeze panes ──────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    for t in range(n_periods):
        ws.column_dimensions[get_column_letter(t + 4)].width = 7

    ws.freeze_panes = "D2"   # freeze unit names + header row

    wb.save(path)
    print(f"  Dispatch matrix exported → {path}\n")


if __name__ == "__main__":
    main()
